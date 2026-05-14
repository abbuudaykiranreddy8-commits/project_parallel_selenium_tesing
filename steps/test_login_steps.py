#code behind file
from pytest_bdd import scenarios, given, when, then, parsers
import time
from pages.login_page import LoginPage
from openpyxl import load_workbook
<<<<<<< HEAD
from utils.db_utils import get_login_data

scenarios("../features/login.feature")

username, password = get_login_data() 
=======

scenarios("../features/login.feature")

>>>>>>> 7920d4db709d59caaf1f1c9f9f5092305e867638
@given("user launches saucedemo site")
def launch_site(driver):
    page = LoginPage(driver)
    page.open()

<<<<<<< HEAD
@when(parsers.parse('user enters username'))
def enter_username(driver):
    page = LoginPage(driver)
    wb=load_workbook(r"C:\Users\Dell\project_parallel_selenium_tesing\sample.xlsx")
    page.enter_username(username)
    
@when(parsers.parse('user enters password'))
def enter_password(driver):
=======
@when(parsers.parse('user enters username "{username}"'))
def enter_username(driver, username):
    page = LoginPage(driver)
    wb=load_workbook(r"C:\Users\Dell\project_parallel_selenium_tesing\sample.xlsx")
    sheet = wb.active
    
    username = sheet["A2"].value
    page.enter_username(username)
    
@when(parsers.parse('user enters password "{password}"'))
def enter_password(driver, password):
>>>>>>> 7920d4db709d59caaf1f1c9f9f5092305e867638
   
    page= LoginPage(driver)
    page.enter_password(password)
   

@when("user clicks login button")
def click_login(driver):
    page= LoginPage(driver)
<<<<<<< HEAD
    time.sleep(30)
=======
    
>>>>>>> 7920d4db709d59caaf1f1c9f9f5092305e867638
    page.click_login()
  
    
@then("user should see products page")
def verify_pdoducts_page(driver):
    page= LoginPage(driver)
    assert "inventory" in driver.current_url
    

        
    
           